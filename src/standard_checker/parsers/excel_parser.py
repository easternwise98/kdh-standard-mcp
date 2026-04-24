import re
from collections import Counter

import openpyxl
from openpyxl.utils import get_column_letter

DEFAULT_COL_WIDTH_CHARS = 8.43
DEFAULT_ROW_HEIGHT_PT = 15.0


def _col_width_to_px(width_chars: float | None) -> int:
    if not width_chars:
        width_chars = DEFAULT_COL_WIDTH_CHARS
    return max(20, int(round(width_chars * 7 + 5)))


def _row_height_to_px(height_pt: float | None) -> int:
    if not height_pt:
        height_pt = DEFAULT_ROW_HEIGHT_PT
    return max(14, int(round(height_pt * 96 / 72)))


MAX_COMPACT_LINES = 80
MAX_COMPACT_CHARS = 4000
KEYWORD_PATTERN = re.compile(
    r"(콘크리트|철근|토압|흙막이|가시설|옹벽|기초|파일|굴착|되메우기|거푸집|배근|슬래브|보|기둥|벽체|"
    r"하중|강도|타설|줄눈|균열|방수|배수|거더|교량|구조|시공|품질|안전|기준|규격)"
)


def _normalize_cell(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    return str(cell).strip()


def _is_formula(value) -> bool:
    return isinstance(value, str) and len(value) > 1 and value.startswith("=")


def _normalize_formula_cell(cached_value, formula_value) -> tuple[str, dict | None]:
    formula = str(formula_value).strip()
    cached_text = _normalize_cell(cached_value)
    if cached_text:
        return cached_text, {"formula": formula, "value": cached_text, "missing_cached_value": False}
    return f"[수식결과없음] {formula}", {"formula": formula, "value": "", "missing_cached_value": True}


def _rewind_if_possible(file) -> None:
    if hasattr(file, "seek"):
        file.seek(0)


def _normalize_row(row: list[str]) -> list[str]:
    cleaned = [re.sub(r"\s+", " ", cell).strip() for cell in row]
    while cleaned and cleaned[-1] == "":
        cleaned.pop()
    return cleaned


def _row_to_text(row: list[str]) -> str:
    return " | ".join(cell for cell in row if cell)


def _has_korean_or_alpha(text: str) -> bool:
    return bool(re.search(r"[A-Za-z가-힣]", text))


def _text_ratio(text: str) -> float:
    if not text:
        return 0.0
    meaningful = len(re.findall(r"[A-Za-z가-힣]", text))
    return meaningful / max(len(text), 1)


def _looks_like_title(row: list[str], text: str) -> bool:
    non_empty = [cell for cell in row if cell]
    if not non_empty:
        return False
    if len(non_empty) <= 2 and len(text) <= 60 and _has_korean_or_alpha(text):
        return True
    if any(token in text for token in ("공사", "기준", "시방", "개요", "특기", "일반사항", "적용", "범위", "재료", "시공")):
        return True
    return False


def _looks_like_header(row: list[str], text: str) -> bool:
    non_empty = [cell for cell in row if cell]
    if len(non_empty) < 2:
        return False
    if len(text) > 120:
        return False
    if _text_ratio(text) < 0.2:
        return False
    short_cells = sum(1 for cell in non_empty if len(cell) <= 20)
    return short_cells >= max(2, len(non_empty) - 1)


def _looks_numeric_heavy(text: str) -> bool:
    compact = text.replace(" ", "")
    if not compact:
        return False
    digits = len(re.findall(r"[0-9]", compact))
    letters = len(re.findall(r"[A-Za-z가-힣]", compact))
    return digits > 0 and digits >= letters * 2


def _score_row(row: list[str], row_index: int) -> tuple[int, str]:
    text = _row_to_text(row)
    if not text:
        return -10, text

    score = 0
    non_empty = [cell for cell in row if cell]
    if row_index < 3:
        score += 3
    if _looks_like_title(row, text):
        score += 6
    if _looks_like_header(row, text):
        score += 4
    if KEYWORD_PATTERN.search(text):
        score += 5
    if any(len(cell) >= 18 for cell in non_empty):
        score += 3
    if any(token in text for token in ("비고", "적용", "조건", "기준", "방법", "주의", "검토", "결과", "설계")):
        score += 2
    if _looks_numeric_heavy(text):
        score -= 3
    if _text_ratio(text) < 0.12:
        score -= 4
    if len(text) < 4:
        score -= 2
    return score, text


def _build_compact_text(rows: list[list[str]]) -> str:
    scored_rows = []
    seen = Counter()

    for row_index, row in enumerate(rows):
        normalized = _normalize_row(row)
        if not normalized:
            continue

        score, text = _score_row(normalized, row_index)
        if not text:
            continue

        seen[text] += 1
        if seen[text] > 2:
            score -= 3

        if score >= 3:
            scored_rows.append((row_index, score, text))

    if not scored_rows:
        fallback = []
        for row in rows[:20]:
            text = _row_to_text(_normalize_row(row))
            if text:
                fallback.append(text)
        compact = "\n".join(fallback)
        return compact[:MAX_COMPACT_CHARS]

    scored_rows.sort(key=lambda item: (-item[1], item[0]))
    selected = scored_rows[:MAX_COMPACT_LINES]
    selected.sort(key=lambda item: item[0])

    compact_lines = []
    total_chars = 0
    for _, _, text in selected:
        line = text[:240]
        next_total = total_chars + len(line) + 1
        if next_total > MAX_COMPACT_CHARS:
            break
        compact_lines.append(line)
        total_chars = next_total

    return "\n".join(compact_lines)


def parse_excel(file) -> list[dict]:
    """
    Parse an Excel workbook sheet-by-sheet.

    Returns a list of dictionaries:
    [
        {
            "sheet": "Sheet1",
            "text": "...raw flattened text...",
            "compact_text": "...LLM-friendly reduced text...",
            "tables": [[...], [...]]
        }
    ]
    """
    result = []
    _rewind_if_possible(file)
    value_workbook = openpyxl.load_workbook(file, data_only=True)
    _rewind_if_possible(file)
    formula_workbook = openpyxl.load_workbook(file, data_only=False)

    for sheet_name in value_workbook.sheetnames:
        worksheet = value_workbook[sheet_name]
        formula_worksheet = formula_workbook[sheet_name]
        rows = []
        kept_row_indices: list[int] = []
        formula_cells: list[dict] = []

        max_row = max(worksheet.max_row, formula_worksheet.max_row)
        max_column = max(worksheet.max_column, formula_worksheet.max_column)

        for excel_row_idx in range(1, max_row + 1):
            normalized = []
            for col_idx in range(1, max_column + 1):
                cached_value = worksheet.cell(excel_row_idx, col_idx).value
                formula_value = formula_worksheet.cell(excel_row_idx, col_idx).value
                if _is_formula(formula_value):
                    text, formula_info = _normalize_formula_cell(cached_value, formula_value)
                    if formula_info is not None:
                        formula_info["cell"] = f"{get_column_letter(col_idx)}{excel_row_idx}"
                        formula_cells.append(formula_info)
                    normalized.append(text)
                else:
                    normalized.append(_normalize_cell(cached_value))
            if any(cell for cell in normalized):
                rows.append(normalized)
                kept_row_indices.append(excel_row_idx)

        max_cols = max((len(r) for r in rows), default=0)
        column_widths_px: list[int] = []
        for i in range(max_cols):
            letter = get_column_letter(i + 1)
            dim = worksheet.column_dimensions.get(letter)
            width_chars = getattr(dim, "width", None) if dim is not None else None
            column_widths_px.append(_col_width_to_px(width_chars))

        row_heights_px: list[int] = []
        for original_idx in kept_row_indices:
            dim = worksheet.row_dimensions.get(original_idx)
            height_pt = getattr(dim, "height", None) if dim is not None else None
            row_heights_px.append(_row_height_to_px(height_pt))

        raw_text = "\n".join(_row_to_text(_normalize_row(row)) for row in rows if _normalize_row(row))
        compact_text = _build_compact_text(rows)

        result.append(
            {
                "sheet": sheet_name,
                "text": raw_text.strip(),
                "compact_text": compact_text.strip(),
                "tables": rows,
                "formula_cells": formula_cells,
                "missing_formula_results": [
                    cell for cell in formula_cells if cell.get("missing_cached_value")
                ],
                "column_widths": column_widths_px,
                "row_heights": row_heights_px,
            }
        )

    return result
